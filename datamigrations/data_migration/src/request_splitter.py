from __future__ import annotations

import argparse
import logging
import re
from collections import defaultdict
from pathlib import Path
from typing import Iterable


REQUEST_PATTERN = re.compile(r"^([A-Z]+)-(\d{4})-(\d+)(?:-([A-Z]+))?$")
LOGGER = logging.getLogger(__name__)


def normalize_request_id(value: str) -> str:
    normalized = re.sub(r"\s*-\s*", "-", value.strip().upper())
    match = REQUEST_PATTERN.fullmatch(normalized)
    if not match:
        raise ValueError(f"Invalid request identifier: {value!r}")
    prefix, year, number, suffix = match.groups()
    parts = [prefix, year, number]
    if suffix:
        parts.append(suffix)
    return "-".join(parts)


def sort_requests(values: Iterable[str]) -> list[str]:
    return sorted(values, key=_sort_key)


def group_requests_by_prefix(values: Iterable[str]) -> dict[str, list[str]]:
    grouped: dict[str, list[str]] = defaultdict(list)
    for value in values:
        normalized = normalize_request_id(value)
        prefix = normalized.split("-", 1)[0]
        grouped[prefix].append(normalized)

    return {prefix: sort_requests(requests) for prefix, requests in sorted(grouped.items())}


def split_request_workbook(input_path: Path, output_dir: Path) -> list[Path]:
    from openpyxl import Workbook, load_workbook

    workbook = load_workbook(input_path)
    sheet = workbook.active
    header = sheet["A1"].value
    if header != "Request":
        raise ValueError("Input workbook must have 'Request' in cell A1.")

    values = []
    for row in range(2, sheet.max_row + 1):
        cell_value = sheet[f"A{row}"].value
        if cell_value is None:
            continue
        text = str(cell_value).strip()
        if not text:
            continue
        try:
            values.append(normalize_request_id(text))
        except ValueError:
            LOGGER.warning("Skipping invalid request identifier %r at row %s", text, row)

    grouped = group_requests_by_prefix(values)
    output_dir.mkdir(parents=True, exist_ok=True)

    output_paths: list[Path] = []
    for prefix, requests in grouped.items():
        output_path = output_dir / f"{prefix}_requests.xlsx"
        output_workbook = Workbook()
        output_sheet = output_workbook.active
        output_sheet.title = "Requests"
        output_sheet["A1"] = "Request"
        for index, request in enumerate(requests, start=2):
            output_sheet[f"A{index}"] = request
        output_workbook.save(output_path)
        output_paths.append(output_path)

    return output_paths


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Split a request workbook into one workbook per prefix.")
    parser.add_argument("--input-xlsx", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    return parser


def run(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    split_request_workbook(args.input_xlsx, args.output_dir)
    return 0


def _sort_key(request_id: str) -> tuple[str, int, int, int, str]:
    prefix, year, number, suffix = _parse_request_id(request_id)
    suffix_rank = 0 if suffix else 1
    return (prefix, int(year), int(number), suffix_rank, suffix or "")


def _parse_request_id(request_id: str) -> tuple[str, str, str, str | None]:
    match = REQUEST_PATTERN.fullmatch(request_id)
    if not match:
        raise ValueError(f"Invalid request identifier: {request_id!r}")
    prefix, year, number, suffix = match.groups()
    return prefix, year, number, suffix


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(run())
