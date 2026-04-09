from pathlib import Path

import pytest

from request_splitter import (
    group_requests_by_prefix,
    normalize_request_id,
    sort_requests,
    split_request_workbook,
)


def test_normalize_request_id_collapses_suffix_spacing() -> None:
    assert normalize_request_id("COR-2025-40987 - DR") == "COR-2025-40987-DR"
    assert normalize_request_id("COR-2025-40987- DR") == "COR-2025-40987-DR"
    assert normalize_request_id("COR-2025-40987 -DR") == "COR-2025-40987-DR"


def test_sort_requests_places_suffix_variants_before_base_request() -> None:
    requests = [
        "COR-2025-40987",
        "COR-2025-40987-R",
        "COR-2025-40987-DR",
        "AGR-2023-30030",
    ]

    assert sort_requests(requests) == [
        "AGR-2023-30030",
        "COR-2025-40987-DR",
        "COR-2025-40987-R",
        "COR-2025-40987",
    ]


def test_group_requests_by_prefix_uses_normalized_values() -> None:
    grouped = group_requests_by_prefix(
        [
            "AGR-2023-30030",
            "COR-2025-40987 - DR",
            "COR-2025-40987",
            "CAF-2023-09887",
        ]
    )

    assert grouped == {
        "AGR": ["AGR-2023-30030"],
        "CAF": ["CAF-2023-09887"],
        "COR": ["COR-2025-40987-DR", "COR-2025-40987"],
    }


def test_split_request_workbook_writes_prefix_files(tmp_path: Path) -> None:
    from openpyxl import Workbook, load_workbook

    input_file = tmp_path / "requests.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Requests"
    sheet["A1"] = "Request"
    sheet["A2"] = "AGR-2023-30030"
    sheet["A3"] = "COR-2025-40987 - DR"
    sheet["A4"] = "COR-2025-40987"
    sheet["A5"] = "CAF-2023-09887"
    workbook.save(input_file)

    output_paths = split_request_workbook(input_file, tmp_path / "out")

    assert [path.name for path in output_paths] == [
        "AGR_requests.xlsx",
        "CAF_requests.xlsx",
        "COR_requests.xlsx",
    ]

    cor_workbook = load_workbook(tmp_path / "out" / "COR_requests.xlsx")
    cor_sheet = cor_workbook.active
    cor_values = [cor_sheet[f"A{row}"].value for row in range(1, cor_sheet.max_row + 1)]

    assert cor_values == ["Request", "COR-2025-40987-DR", "COR-2025-40987"]


def test_split_request_workbook_logs_and_skips_invalid_requests(tmp_path: Path, caplog: pytest.LogCaptureFixture) -> None:
    from openpyxl import Workbook, load_workbook

    input_file = tmp_path / "requests.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Request"
    sheet["A2"] = "[REQUESTNUMBER]"
    sheet["A3"] = "COR-2025-40987 - DR"
    workbook.save(input_file)

    with caplog.at_level("WARNING"):
        output_paths = split_request_workbook(input_file, tmp_path / "out")

    assert [path.name for path in output_paths] == ["COR_requests.xlsx"]
    assert "Skipping invalid request identifier '[REQUESTNUMBER]'" in caplog.text

    cor_workbook = load_workbook(tmp_path / "out" / "COR_requests.xlsx")
    cor_sheet = cor_workbook.active
    cor_values = [cor_sheet[f"A{row}"].value for row in range(1, cor_sheet.max_row + 1)]

    assert cor_values == ["Request", "COR-2025-40987-DR"]
